from __future__ import annotations

import csv
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from base_builder import build_management_workbook
from main import process_text_file


class IntegrationWorkflowTests(unittest.TestCase):
    def _write_csv(self, path: Path, rows: list[dict]) -> None:
        path.parent.mkdir(parents=True, exist_ok=True)
        with path.open("w", encoding="utf-8-sig", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
            writer.writeheader()
            writer.writerows(rows)

    def test_fluxo_completo_gera_base_gerencial_sem_quebrar_layout(self):
        mensagem = """
RDO - 11/03/2026

NUCLEO: Mississipi
LOGRADOURO: Viela 04 - acesso pela Rua Sao Jorge
MUNICIPIO: Carapicuiba
EQUIPE: Equipe 03

EXECUCAO:
- 12 un hidrometro
- 3 un servico ultradesconhecido xyz

FRENTES:
- assentamento de rede

OCORRENCIAS:
- interferencia com rede de esgoto

OBS:
- trecho com acesso estreito
""".strip()

        with tempfile.TemporaryDirectory() as tmp_dir:
            tmp_path = Path(tmp_dir)
            input_path = tmp_path / "entrada_oficial.txt"
            output_dir = tmp_path / "saida"
            input_path.write_text(mensagem + "\n", encoding="utf-8")

            process_text_file(input_path, output_dir)

            execucao_csv = output_dir / "execucao.csv"
            frentes_csv = output_dir / "frentes.csv"
            ocorrencias_csv = output_dir / "ocorrencias.csv"
            nao_mapeado_csv = output_dir / "servico_nao_mapeado.csv"
            workbook_path = output_dir / "base_gerencial.xlsx"

            self.assertTrue(execucao_csv.exists())
            self.assertTrue(frentes_csv.exists())
            self.assertTrue(ocorrencias_csv.exists())
            self.assertTrue(nao_mapeado_csv.exists())
            self.assertTrue(workbook_path.exists())

            with execucao_csv.open("r", encoding="utf-8-sig", newline="") as f:
                reader = csv.DictReader(f)
                headers = reader.fieldnames or []
                rows = list(reader)
            for col in ["servico_bruto", "servico_normalizado", "servico_oficial", "categoria"]:
                self.assertIn(col, headers)
            self.assertGreaterEqual(len(rows), 2)

            with nao_mapeado_csv.open("r", encoding="utf-8-sig", newline="") as f:
                reader = csv.DictReader(f)
                nao_mapeados = list(reader)
            self.assertGreaterEqual(len(nao_mapeados), 1)

            wb = load_workbook(workbook_path)
            self.assertIn("DASHBOARD_EXECUTIVO", wb.sheetnames)
            ws_dash = wb["DASHBOARD_EXECUTIVO"]
            titulo = str(ws_dash["B2"].value or "")
            self.assertIn("SABESP", titulo.upper())
            self.assertIn("PAINEL GERENCIAL", titulo.upper())

            expected_merges = [
                "B2:O3",
                "B4:O4",
                "B5:O5",
                "B7:D9",
                "G7:I9",
                "L7:N9",
                "B11:D13",
                "G11:I13",
                "L11:N13",
                "B15:O15",
                "B17:F17",
                "H17:K17",
                "M17:O17",
                "B30:E30",
                "H30:I30",
            ]
            self.assertEqual(
                sorted(str(r) for r in ws_dash.merged_cells.ranges),
                sorted(expected_merges),
            )

            expected_widths = {
                "A": 2,
                "B": 16,
                "C": 12,
                "D": 12,
                "E": 12,
                "F": 2,
                "G": 16,
                "H": 12,
                "I": 12,
                "J": 12,
                "K": 2,
                "L": 16,
                "M": 12,
                "N": 12,
                "O": 12,
                "P": 2,
                "Q": 12,
                "R": 12,
            }
            for col, width in expected_widths.items():
                self.assertEqual(ws_dash.column_dimensions[col].width, width)

            self.assertFalse(str(ws_dash["J31"].value or "").endswith(","))
            self.assertFalse(str(ws_dash["J32"].value or "").endswith(","))
            self.assertFalse(str(ws_dash["F19"].value or "").endswith(","))

            self.assertIn("VISAO_NUCLEOS", wb.sheetnames)
            ws_nuc = wb["VISAO_NUCLEOS"]
            self.assertFalse(str(ws_nuc["B4"].value or "").endswith(","))
            self.assertFalse(str(ws_nuc["E4"].value or "").endswith(","))

            self.assertIn("DICIONARIO_SERVICOS", wb.sheetnames)
            ws_dic = wb["DICIONARIO_SERVICOS"]
            dic_sample = " ".join(
                str(ws_dic[f"{col}{row}"].value or "")
                for row in range(4, 14)
                for col in ("A", "B", "F", "G")
            )
            for token in ("Ãƒ", "Ã‚", "Ã¢", "ï¿½"):
                self.assertNotIn(token, dic_sample)
            self.assertIn("Execução", dic_sample)

            wb.close()

    def test_processamento_persiste_apenas_primeira_equipe_nos_csvs(self):
        mensagem = """
RDO - 11/03/2026

NUCLEO: Mississipi
LOGRADOURO: Viela 04
MUNICIPIO: Carapicuiba
EQUIPE: Equipe 01 / Equipe 02

EXECUCAO:
- 2 un hidrometro

FRENTES:
- frente unica

OCORRENCIAS:
- sem ocorrencias
""".strip()

        with tempfile.TemporaryDirectory() as tmp_dir:
            tmp_path = Path(tmp_dir)
            input_path = tmp_path / "entrada_equipe.txt"
            output_dir = tmp_path / "saida"
            input_path.write_text(mensagem + "\n", encoding="utf-8")

            process_text_file(input_path, output_dir)

            for name in ("execucao.csv", "frentes.csv", "ocorrencias.csv", "observacoes.csv"):
                with (output_dir / name).open("r", encoding="utf-8-sig", newline="") as f:
                    rows = list(csv.DictReader(f))
                if name != "observacoes.csv":
                    self.assertTrue(rows)
                self.assertTrue(all(str(r.get("equipe", "") or "").strip() == "Equipe 01" for r in rows))

    def test_processamento_cli_reconcilia_municipio_oficial_por_nucleo(self):
        mensagem = """
RDO - 11/03/2026

NUCLEO: Patrocinio
LOGRADOURO: Rua 1
MUNICIPIO: Barueri
EQUIPE: Equipe 01

EXECUCAO:
- 1 un hidrometro
""".strip()

        with tempfile.TemporaryDirectory() as tmp_dir:
            tmp_path = Path(tmp_dir)
            input_path = tmp_path / "entrada_oficial_nucleo.txt"
            output_dir = tmp_path / "saida"
            input_path.write_text(mensagem + "\n", encoding="utf-8")

            process_text_file(input_path, output_dir)

            with (output_dir / "execucao.csv").open("r", encoding="utf-8-sig", newline="") as f:
                rows = list(csv.DictReader(f))

            self.assertTrue(rows)
            self.assertEqual("Patrocínio", rows[0]["nucleo"])
            self.assertEqual("Carapicuíba", rows[0]["municipio"])
            self.assertEqual("Patrocinio", rows[0]["nucleo_detectado_texto"])
            self.assertEqual("Barueri", rows[0]["municipio_detectado_texto"])
            self.assertEqual("Patrocínio", rows[0]["nucleo_oficial"])
            self.assertEqual("Carapicuíba", rows[0]["municipio_oficial"])

    def test_caixa_uma_total_nao_duplica_embutida_mureta(self):
        mensagem = """
RDO - AL 33 - Oeste 1
27/02/2026

Nucleo: Mississipi (Equipe Paulo - Viela 1)
Caixas UMA: 16
Embutida: 7
Mureta: 9
""".strip()

        with tempfile.TemporaryDirectory() as tmp_dir:
            tmp_path = Path(tmp_dir)
            input_path = tmp_path / "entrada_caixa_uma.txt"
            output_dir = tmp_path / "saida"
            input_path.write_text(mensagem + "\n", encoding="utf-8")

            process_text_file(input_path, output_dir)

            painel_csv = output_dir / "painel_geral.csv"
            indicadores_csv = output_dir / "indicadores_dashboard.csv"
            self.assertTrue(painel_csv.exists())
            self.assertTrue(indicadores_csv.exists())

            with painel_csv.open("r", encoding="utf-8-sig", newline="") as f:
                painel_rows = list(csv.DictReader(f))
            self.assertEqual(len(painel_rows), 1)
            self.assertEqual(float(painel_rows[0].get("quantidade_total", 0) or 0), 16.0)
            self.assertEqual(float(painel_rows[0].get("itens_total", 0) or 0), 1.0)

            with indicadores_csv.open("r", encoding="utf-8-sig", newline="") as f:
                ind_rows = list(csv.DictReader(f))
            caixa_rows = [r for r in ind_rows if str(r.get("categoria_item", "") or "").strip() == "caixa_uma"]
            self.assertGreaterEqual(len(caixa_rows), 1)
            total_qtd = sum(float(r.get("quantidade_total", 0) or 0) for r in caixa_rows)
            total_reg = sum(float(r.get("registros", 0) or 0) for r in caixa_rows)
            self.assertEqual(total_qtd, 16.0)
            self.assertEqual(total_reg, 1.0)

    def test_reprocessamento_na_mesma_saida_limpa_csv_vazio(self):
        mensagem_1 = """
RDO - 11/03/2026

NUCLEO: Mississipi
LOGRADOURO: Rua 1
MUNICIPIO: Carapicuiba
EQUIPE: Equipe 01

EXECUCAO:
- 1 un hidrometro

FRENTES:
- frente 1

OCORRENCIAS:
- chuva
""".strip()

        mensagem_2 = """
RDO - 12/03/2026

NUCLEO: Mississipi
LOGRADOURO: Rua 1
MUNICIPIO: Carapicuiba
EQUIPE: Equipe 01

EXECUCAO:
- 1 un hidrometro

FRENTES:
- frente 1

OCORRENCIAS:
""".strip()

        with tempfile.TemporaryDirectory() as tmp_dir:
            tmp_path = Path(tmp_dir)
            output_dir = tmp_path / "saida"
            p1 = tmp_path / "entrada_1.txt"
            p2 = tmp_path / "entrada_2.txt"
            p1.write_text(mensagem_1 + "\n", encoding="utf-8")
            p2.write_text(mensagem_2 + "\n", encoding="utf-8")

            process_text_file(p1, output_dir)
            process_text_file(p2, output_dir)

            ocorrencias_csv = output_dir / "ocorrencias.csv"
            self.assertTrue(ocorrencias_csv.exists())

            with ocorrencias_csv.open("r", encoding="utf-8-sig", newline="") as f:
                rows = list(csv.DictReader(f))
            self.assertEqual(rows, [])

    def test_dashboard_consolida_nucleo_sem_replicar_municipios_nas_agregacoes(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            output_dir = Path(tmp_dir) / "saida"
            output_dir.mkdir(parents=True, exist_ok=True)

            self._write_csv(
                output_dir / "frentes.csv",
                [
                    {
                        "id_frente": "F001",
                        "data_referencia": "12/03/2026",
                        "nucleo": "Centro",
                        "municipio": "Jandira",
                        "equipe": "Equipe A",
                        "logradouro": "Rua A",
                        "status_frente": "ativa",
                    },
                    {
                        "id_frente": "F002",
                        "data_referencia": "12/03/2026",
                        "nucleo": "Centro",
                        "municipio": "Carapicuiba",
                        "equipe": "Equipe B",
                        "logradouro": "Rua B",
                        "status_frente": "ativa",
                    },
                ],
            )

            self._write_csv(
                output_dir / "execucao.csv",
                [
                    {
                        "id_item": "I001",
                        "id_frente": "F001",
                        "data_referencia": "12/03/2026",
                        "nucleo": "Centro",
                        "municipio": "Jandira",
                        "equipe": "Equipe A",
                        "categoria_item": "rede_agua",
                        "quantidade": "10",
                    },
                    {
                        "id_item": "I002",
                        "id_frente": "F002",
                        "data_referencia": "12/03/2026",
                        "nucleo": "Centro",
                        "municipio": "Carapicuiba",
                        "equipe": "Equipe B",
                        "categoria_item": "rede_agua",
                        "quantidade": "12",
                    },
                ],
            )

            self._write_csv(
                output_dir / "ocorrencias.csv",
                [
                    {
                        "id_ocorrencia": "O001",
                        "id_frente": "F001",
                        "data_referencia": "12/03/2026",
                        "nucleo": "Centro",
                        "municipio": "Jandira",
                        "tipo_ocorrencia": "interferencia",
                        "descricao": "rede existente",
                    }
                ],
            )

            wb_path = build_management_workbook(output_dir)
            wb = load_workbook(wb_path)
            ws_nuc = wb["VISAO_NUCLEOS"]

            labels = {str(ws_nuc[f"A{row}"].value or "").strip() for row in range(4, 16)}
            labels.discard("")
            self.assertIn("Centro", labels)
            self.assertFalse(any(label.startswith("Centro (") for label in labels))

            wb.close()

    def test_dashboard_colapsa_variacoes_de_mississipi_e_ignora_vila_dirce(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            output_dir = Path(tmp_dir) / "saida"
            output_dir.mkdir(parents=True, exist_ok=True)

            self._write_csv(
                output_dir / "frentes.csv",
                [
                    {
                        "id_frente": "F001",
                        "data_referencia": "12/03/2026",
                        "nucleo": "Mississipi - Caixa UMA",
                        "municipio": "Barueri",
                        "equipe": "Equipe A",
                        "logradouro": "Viela 1",
                        "status_frente": "ativa",
                    },
                    {
                        "id_frente": "F002",
                        "data_referencia": "12/03/2026",
                        "nucleo": "Vila Dirce",
                        "municipio": "Carapicuiba",
                        "equipe": "Equipe B",
                        "logradouro": "Rua B",
                        "status_frente": "ativa",
                    },
                ],
            )

            self._write_csv(
                output_dir / "execucao.csv",
                [
                    {
                        "id_item": "I001",
                        "id_frente": "F001",
                        "data_referencia": "12/03/2026",
                        "nucleo": "Mississipi - Esgoto",
                        "municipio": "Barueri",
                        "equipe": "Equipe A",
                        "categoria_item": "rede_agua",
                        "quantidade": "10",
                    },
                    {
                        "id_item": "I002",
                        "id_frente": "F001",
                        "data_referencia": "12/03/2026",
                        "nucleo": "Viela 1 / Viela 7",
                        "municipio": "Barueri",
                        "equipe": "Equipe A",
                        "categoria_item": "rede_agua",
                        "quantidade": "12",
                    },
                    {
                        "id_item": "I003",
                        "id_frente": "F002",
                        "data_referencia": "12/03/2026",
                        "nucleo": "Vila Dirce",
                        "municipio": "Carapicuiba",
                        "equipe": "Equipe B",
                        "categoria_item": "rede_agua",
                        "quantidade": "8",
                    },
                ],
            )

            self._write_csv(
                output_dir / "ocorrencias.csv",
                [
                    {
                        "id_ocorrencia": "O001",
                        "id_frente": "F001",
                        "data_referencia": "12/03/2026",
                        "nucleo": "Mississipi - Esgoto",
                        "municipio": "Barueri",
                        "tipo_ocorrencia": "interferencia",
                        "descricao": "rede existente",
                    }
                ],
            )

            wb_path = build_management_workbook(output_dir)
            wb = load_workbook(wb_path)
            ws_nuc = wb["VISAO_NUCLEOS"]

            labels = {str(ws_nuc[f"A{row}"].value or "").strip() for row in range(4, 16)}
            labels.discard("")
            self.assertIn("Mississipi", labels)
            self.assertNotIn("Mississipi - Caixa UMA", labels)
            self.assertNotIn("Mississipi - Esgoto", labels)
            self.assertNotIn("Viela 1 / Viela 7", labels)
            self.assertNotIn("Vila Dirce", labels)

            wb.close()


    def test_relatorio_md_sem_mojibake(self):
        mensagem = """
RDO - 14/03/2026

NUCLEO: Cerejas
LOGRADOURO: Travessa das Cerejas
MUNICIPIO: Carapicuiba
EQUIPE: Equipe 07

EXECUCAO:
- 4 un hidrometro
- 10 m rede de agua

FRENTES:
- frente principal

OCORRENCIAS:
- chuva moderada no periodo da tarde
""".strip()

        with tempfile.TemporaryDirectory() as tmp_dir:
            tmp_path = Path(tmp_dir)
            input_path = tmp_path / "entrada_relatorio.txt"
            output_dir = tmp_path / "saida"
            input_path.write_text(mensagem + "\n", encoding="utf-8")

            process_text_file(input_path, output_dir)

            relatorios_dir = output_dir / "relatorios_nucleos"
            self.assertTrue(relatorios_dir.exists())
            md_files = list(relatorios_dir.glob("*.md"))
            self.assertGreaterEqual(len(md_files), 1)

            conteudo = md_files[0].read_text(encoding="utf-8")
            self.assertIn("RELATÓRIO TÉCNICO DE EVOLUÇÃO DE OBRA", conteudo)
            for token in ("NÃƒ", "ÃƒÂ§", "ÃƒÂ£", "ÃƒÂ¡", "ÃƒÂ©", "ÃƒÂ­", "ÃƒÂ³", "ÃƒÂº", "Ã¢â‚¬Â¢", "Ã¢â‚¬â€œ", "Ã‚"):
                self.assertNotIn(token, conteudo)

if __name__ == "__main__":
    unittest.main()








