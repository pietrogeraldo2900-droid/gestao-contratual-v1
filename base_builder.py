from __future__ import annotations

import csv
import unicodedata
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from nucleo_master import get_nucleo_profile, load_nucleo_registry, reconcile_rows_with_registry

# Institutional palette
NAVY = "0A3A66"
BLUE = "0E5E9E"
LIGHT_BLUE = "DCEAF8"
VERY_LIGHT = "F4F8FC"
WHITE = "FFFFFF"
LINE = "D3DEE9"
TEXT = "1F2937"
MUTED = "5B6B7C"
SUCCESS_BG = "EAF6EC"
WARNING_BG = "FFF4E6"
DANGER_BG = "FDECEC"
NEUTRAL_BG = "EEF2F7"

BASE_DIR = Path(__file__).resolve().parent
NUCLEO_REFERENCE_FILE = BASE_DIR / "config" / "nucleo_reference.json"

THIN = Border(
    left=Side(style="thin", color=LINE),
    right=Side(style="thin", color=LINE),
    top=Side(style="thin", color=LINE),
    bottom=Side(style="thin", color=LINE),
)


def _repair_mojibake(value: object) -> str:
    text = str(value or "")
    suspicious_chars = (chr(0x00C3), chr(0x00C2), chr(0x00E2), chr(0x00F0), chr(0xFFFD))
    if any(ch in text for ch in suspicious_chars):
        try:
            fixed = text.encode("latin1", errors="ignore").decode("utf-8", errors="ignore")
            if fixed:
                return fixed
        except Exception:
            pass
    return text


def _read_csv(path: Path) -> List[dict]:
    if not path.exists():
        return []
    for enc in ("utf-8-sig", "utf-8", "cp1252", "latin1"):
        try:
            with path.open("r", encoding=enc, newline="") as f:
                rows = list(csv.DictReader(f))
            return [{k: _repair_mojibake(v) for k, v in row.items()} for row in rows]
        except Exception:
            continue
    return []


def _reconcile_rows(rows: List[dict], nucleo_reference_file: Path | None = None) -> List[dict]:
    registry = load_nucleo_registry(nucleo_reference_file or NUCLEO_REFERENCE_FILE)
    return reconcile_rows_with_registry(rows, registry)


def _filter_active_nucleo_rows(
    rows: List[dict], nucleo_reference_file: Path | None = None
) -> List[dict]:
    registry = load_nucleo_registry(nucleo_reference_file or NUCLEO_REFERENCE_FILE)
    filtered: List[dict] = []
    for row in rows:
        nucleo = str(row.get("nucleo", "") or "").strip()
        if not nucleo:
            filtered.append(row)
            continue
        profile = get_nucleo_profile(registry, nucleo)
        if profile and str(profile.get("status", "") or "").strip().lower() == "inativo":
            continue
        filtered.append(row)
    return filtered

def _num(value) -> float:
    if value in (None, ""):
        return 0.0
    s = str(value).strip().replace(" ", "")
    if not s:
        return 0.0
    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    else:
        s = s.replace(",", ".")
    try:
        return float(s)
    except Exception:
        return 0.0


def _maybe_int(value: float):
    try:
        v = float(value)
    except Exception:
        return value
    if v.is_integer():
        return int(v)
    return round(v, 2)


def _normalize_text(value: object) -> str:
    text = str(value or "").strip().lower()
    return "".join(ch for ch in unicodedata.normalize("NFKD", text) if not unicodedata.combining(ch))


def _first_non_empty(*values: str) -> str:
    for value in values:
        txt = str(value or "").strip()
        if txt:
            return txt
    return ""


def _style_cell(
    cell,
    *,
    fill: str = WHITE,
    bold: bool = False,
    color: str = TEXT,
    size: int = 10,
    align: str = "left",
    valign: str = "center",
    wrap: bool = False,
) -> None:
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.font = Font(name="Calibri", size=size, bold=bold, color=color)
    cell.alignment = Alignment(horizontal=align, vertical=valign, wrap_text=wrap)
    cell.border = THIN


def _merge_box(
    ws,
    start_row: int,
    start_col: int,
    end_row: int,
    end_col: int,
    text: str,
    *,
    fill: str,
    color: str,
    size: int,
    align: str = "center",
    bold: bool = True,
) -> None:
    ws.merge_cells(start_row=start_row, start_column=start_col, end_row=end_row, end_column=end_col)
    top_left = ws.cell(start_row, start_col)
    top_left.value = text
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            _style_cell(ws.cell(r, c), fill=fill, color=color, size=size, align=align, bold=bold, wrap=True)


def _kpi_card(ws, start_row: int, start_col: int, title: str, value, subtitle: str) -> None:
    end_row = start_row + 2
    end_col = start_col + 2
    text = f"{title}\n{value}\n{subtitle}"
    _merge_box(
        ws,
        start_row,
        start_col,
        end_row,
        end_col,
        text,
        fill=LIGHT_BLUE,
        color=NAVY,
        size=13,
        align="center",
        bold=True,
    )


def _write_table(
    ws,
    start_row: int,
    start_col: int,
    headers: List[str],
    rows: List[List[object]],
    *,
    max_rows: int,
    status_col_idx: int | None = None,
) -> None:
    for i, h in enumerate(headers):
        c = ws.cell(start_row, start_col + i, h)
        _style_cell(c, fill=NAVY, color=WHITE, bold=True, align="center")

    for i in range(max_rows):
        row_idx = start_row + 1 + i
        row_data = rows[i] if i < len(rows) else [""] * len(headers)
        for j, value in enumerate(row_data):
            cell = ws.cell(row_idx, start_col + j)
            stripe = VERY_LIGHT if i % 2 == 0 else WHITE
            if isinstance(value, bool):
                value = int(value)
            if isinstance(value, float):
                cell_value = _maybe_int(value)
            elif isinstance(value, int):
                cell_value = value
            else:
                cell_value = value
            cell.value = cell_value

            numeric = isinstance(cell_value, (int, float))
            alignment = "right" if numeric else "left"
            _style_cell(cell, fill=stripe, color=TEXT, align=alignment)

            if isinstance(cell_value, int):
                cell.number_format = "0"
            elif isinstance(cell_value, float):
                cell.number_format = "0.##"

        if status_col_idx is not None:
            status_cell = ws.cell(row_idx, start_col + status_col_idx)
            _style_status_cell(status_cell)


def _style_status_cell(cell) -> None:
    raw = _normalize_text(cell.value)
    if not raw:
        fill = NEUTRAL_BG
    elif any(token in raw for token in ("ativa", "ativo", "estavel", "destaque", "baixo", "ok")):
        fill = SUCCESS_BG
    elif any(token in raw for token in ("monitorar", "atencao", "moderado", "parcial", "medio")):
        fill = WARNING_BG
    elif any(token in raw for token in ("alto", "critico", "sem", "alerta")):
        fill = DANGER_BG
    else:
        fill = NEUTRAL_BG
    _style_cell(cell, fill=fill, color=TEXT, bold=True, align="center")


def _autosize(ws, max_col: int) -> None:
    for col in range(1, max_col + 1):
        width = 12
        for row in range(1, min(ws.max_row, 800) + 1):
            value = ws.cell(row, col).value
            if value is None:
                continue
            width = max(width, min(56, len(str(value)) + 2))
        ws.column_dimensions[get_column_letter(col)].width = width


def _normalize_sheet_rows(rows: List[dict], numeric_keys: set[str]) -> List[dict]:
    out: List[dict] = []
    for row in rows:
        copy_row = dict(row)
        for k in numeric_keys:
            if k in copy_row:
                v = _num(copy_row.get(k))
                copy_row[k] = _maybe_int(v)
        out.append(copy_row)
    return out


def _dump_base_sheet(ws, title: str, rows: List[dict], fallback_headers: List[str]) -> None:
    ws.sheet_view.showGridLines = True
    ws.freeze_panes = "A4"

    headers = list(rows[0].keys()) if rows else fallback_headers
    end_col = max(3, len(headers))
    _merge_box(
        ws,
        1,
        1,
        1,
        end_col,
        title,
        fill=NAVY,
        color=WHITE,
        size=13,
        align="left",
    )

    for i, h in enumerate(headers, start=1):
        _style_cell(ws.cell(3, i, h), fill=BLUE, color=WHITE, bold=True, align="center")

    numeric_keys = {
        "quantidade",
        "quantidade_total",
        "registros",
        "frentes",
        "itens",
        "ocorrencias",
        "frentes_total",
        "itens_total",
        "ocorrencias_total",
        "frentes_sem_producao",
    }
    normalized = _normalize_sheet_rows(rows, numeric_keys)

    if normalized:
        for r_idx, row in enumerate(normalized, start=4):
            stripe = VERY_LIGHT if (r_idx - 4) % 2 == 0 else WHITE
            for c_idx, h in enumerate(headers, start=1):
                val = row.get(h, "")
                cell = ws.cell(r_idx, c_idx, val)
                _style_cell(cell, fill=stripe, color=TEXT, align="left")
                if isinstance(val, int):
                    cell.number_format = "0"
                elif isinstance(val, float):
                    cell.number_format = "0.##"
        ws.auto_filter.ref = f"A3:{get_column_letter(len(headers))}{3 + len(normalized)}"
    else:
        _style_cell(ws.cell(4, 1, "Sem dados"), fill=VERY_LIGHT, color=MUTED)

    _autosize(ws, len(headers))


def _exec_scope_key(row: dict) -> str:
    id_frente = str(row.get("id_frente", "") or "").strip()
    if id_frente:
        return f"id:{id_frente}"
    return "ctx:{0}|{1}|{2}|{3}".format(
        str(row.get("data_referencia", "") or "").strip(),
        str(row.get("nucleo", "") or "").strip(),
        str(row.get("equipe", "") or "").strip(),
        str(row.get("logradouro", "") or "").strip(),
    )


def _service_text(row: dict) -> str:
    raw = row.get("servico_bruto") or row.get("item_original") or row.get("item_normalizado")
    return _normalize_text(raw)


def _category_text(row: dict) -> str:
    raw = row.get("categoria_item") or row.get("categoria")
    return _normalize_text(raw).replace(" ", "_")


def _is_caixa_uma_total(row: dict) -> bool:
    if _category_text(row) != "caixa_uma":
        return False
    service = _service_text(row)
    if not service:
        return False
    return ("caixa" in service and "uma" in service) and ("embutid" not in service and "mureta" not in service and "reboco" not in service)


def _is_caixa_uma_breakdown(row: dict) -> bool:
    if _category_text(row) != "caixa_uma":
        return False
    service = _service_text(row)
    return any(token in service for token in ("embutid", "mureta", "reboco"))


def _countable_exec_rows(exec_rows: List[dict]) -> List[dict]:
    total_scopes = {_exec_scope_key(r) for r in exec_rows if _is_caixa_uma_total(r)}
    out: List[dict] = []
    for row in exec_rows:
        if _is_caixa_uma_breakdown(row) and _exec_scope_key(row) in total_scopes:
            continue
        out.append(row)
    return out


def _build_dashboard(
    ws,
    *,
    totals: dict,
    by_nucleus: Dict[str, dict],
    by_category: Dict[str, dict],
    by_team: Dict[str, dict],
    by_occ_type: Counter,
    frentes_sem_producao: int,
    ocorrencia_por_frente: float,
    risco_operacional: str,
    data_base: str,
    data_geracao: str,
) -> None:
    ws.sheet_view.showGridLines = False
    # Column grid A..R
    widths = {
        "A": 2,
        "B": 16,
        "C": 12,
        "D": 12,
        "E": 12,
        "F": 2,
        "G": 16,
        "H": 12,
        "I": 12,
        "J": 12,
        "K": 2,
        "L": 16,
        "M": 12,
        "N": 12,
        "O": 12,
        "P": 2,
        "Q": 12,
        "R": 12,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    for row in range(1, 60):
        ws.row_dimensions[row].height = 21
    ws.row_dimensions[2].height = 26
    ws.row_dimensions[3].height = 26
    ws.row_dimensions[4].height = 18
    ws.row_dimensions[5].height = 22
    ws.row_dimensions[17].height = 24
    ws.row_dimensions[30].height = 24

    _merge_box(
        ws,
        2,
        2,
        3,
        15,
        "SABESP - PAINEL GERENCIAL DE EVOLUCAO DE OBRA",
        fill=NAVY,
        color=WHITE,
        size=17,
        align="center",
    )

    _merge_box(
        ws,
        4,
        2,
        4,
        15,
        "Base gerada automaticamente a partir dos CSVs. Indicadores para leitura executiva.",
        fill=WHITE,
        color=MUTED,
        size=10,
        align="left",
        bold=False,
    )

    meta = (
        f"Data base: {data_base or '-'} | Gerado em: {data_geracao} | "
        f"Frentes sem producao: {frentes_sem_producao} | Risco operacional: {risco_operacional}"
    )
    _merge_box(ws, 5, 2, 5, 15, meta, fill=VERY_LIGHT, color=NAVY, size=10, align="left")

    _kpi_card(ws, 7, 2, "NUCLEOS", totals["nucleos"], "com atividade no periodo")
    _kpi_card(ws, 7, 7, "FRENTES", totals["frentes"], "frentes registradas")
    _kpi_card(ws, 7, 12, "OCORRENCIAS", totals["ocorrencias"], "itens de risco")

    _kpi_card(ws, 11, 2, "QTD TOTAL", totals["quantidade_total"], "volume consolidado")
    _kpi_card(ws, 11, 7, "CATEGORIAS", totals["categorias_ativas"], "categorias ativas")
    _kpi_card(ws, 11, 12, "EQUIPES", totals["equipes_ativas"], "equipes com producao")

    radar = f"RADAR OPERACIONAL | Ocorrencias por frente: {ocorrencia_por_frente} | Nivel de risco: {risco_operacional}"
    _merge_box(ws, 15, 2, 15, 15, radar, fill=VERY_LIGHT, color=NAVY, size=10, align="left")

    _merge_box(ws, 17, 2, 17, 6, "VISAO POR NUCLEO", fill=BLUE, color=WHITE, size=11, align="center")
    _merge_box(ws, 17, 8, 17, 11, "VISAO POR CATEGORIA", fill=BLUE, color=WHITE, size=11, align="center")
    _merge_box(ws, 17, 13, 17, 15, "VISAO POR EQUIPE", fill=BLUE, color=WHITE, size=11, align="center")

    nucleus_rows = []
    for n, v in sorted(by_nucleus.items()):
        nucleus_rows.append([
            n,
            _maybe_int(v["frentes"]),
            _maybe_int(v["itens"]),
            _maybe_int(v["ocorrencias"]),
            _maybe_int(v["quantidade_total"]),
        ])

    category_rows = []
    for c, v in sorted(by_category.items()):
        category_rows.append([
            c,
            _maybe_int(v["registros"]),
            _maybe_int(v["quantidade_total"]),
            "Ativa" if v["registros"] > 0 else "Sem movimento",
        ])

    team_rows = []
    for t, v in sorted(by_team.items()):
        team_rows.append([
            t,
            v["nucleo"],
            _maybe_int(v["quantidade_total"]),
        ])

    _write_table(ws, 18, 2, ["Nucleo", "Frentes", "Itens", "Ocorr.", "Qtd Total"], nucleus_rows, max_rows=8)
    _write_table(ws, 18, 8, ["Categoria", "Registros", "Qtd Total", "Status"], category_rows, max_rows=8, status_col_idx=3)
    _write_table(ws, 18, 13, ["Equipe", "Nucleo", "Qtd Total"], team_rows, max_rows=8)

    _merge_box(ws, 30, 2, 30, 5, "OCORRENCIAS POR TIPO", fill=BLUE, color=WHITE, size=11)
    _merge_box(ws, 30, 8, 30, 9, "LEITURA GERENCIAL", fill=BLUE, color=WHITE, size=11)

    occ_rows = []
    for tp, qtd in sorted(by_occ_type.items()):
        occ_rows.append([
            tp,
            _maybe_int(qtd),
            "Monitorar" if qtd > 0 else "Sem impacto",
            "Ativo" if qtd > 0 else "Sem registro",
        ])

    leitura_rows = [
        ["Frentes cadastradas", _maybe_int(totals["frentes"])],
        ["Itens executados", _maybe_int(totals.get("itens_execucao", 0))],
        ["Ocorrencias registradas", _maybe_int(totals["ocorrencias"])],
        ["Nucleos com producao", _maybe_int(totals["nucleos"])],
        ["Maior categoria em volume", max(by_category.items(), key=lambda x: x[1]["quantidade_total"])[0] if by_category else "-"],
        ["Maior nucleo em volume", max(by_nucleus.items(), key=lambda x: x[1]["quantidade_total"])[0] if by_nucleus else "-"],
        ["Frentes sem producao", _maybe_int(frentes_sem_producao)],
        ["Risco operacional", risco_operacional],
    ]

    _write_table(ws, 31, 2, ["Tipo", "Qtd", "Leitura", "Status"], occ_rows, max_rows=8, status_col_idx=3)
    _write_table(ws, 31, 8, ["Indicador", "Valor"], leitura_rows, max_rows=8, status_col_idx=1)


def _build_view_nuc(ws, by_nucleus: Dict[str, dict]) -> None:
    ws.sheet_view.showGridLines = False
    _merge_box(ws, 1, 1, 1, 6, "SABESP - VISAO POR NUCLEO", fill=NAVY, color=WHITE, size=14, align="left")

    rows = []
    for n, v in sorted(by_nucleus.items()):
        taxa = v["ocorrencias"] / max(1, v["frentes"])
        if v["itens"] == 0:
            comentario = "Sem producao"
        elif taxa >= 1:
            comentario = "Atencao operacional"
        else:
            comentario = "Operacao estavel"
        rows.append([
            n,
            _maybe_int(v["frentes"]),
            _maybe_int(v["itens"]),
            _maybe_int(v["ocorrencias"]),
            _maybe_int(v["quantidade_total"]),
            comentario,
        ])

    _write_table(ws, 3, 1, ["Nucleo", "Frentes", "Itens", "Ocorrencias", "Qtd Total", "Comentario"], rows, max_rows=max(len(rows), 12), status_col_idx=5)
    ws.freeze_panes = "A4"
    if rows:
        ws.auto_filter.ref = f"A3:F{3 + len(rows)}"
    _autosize(ws, 6)


def _build_view_cat(ws, by_category: Dict[str, dict]) -> None:
    ws.sheet_view.showGridLines = False
    _merge_box(ws, 1, 1, 1, 5, "SABESP - VISAO POR CATEGORIA", fill=NAVY, color=WHITE, size=14, align="left")

    total_qty = sum(v["quantidade_total"] for v in by_category.values()) or 1
    rows = []
    for c, v in sorted(by_category.items()):
        part = f"{round((v['quantidade_total'] / total_qty) * 100, 1)}%"
        rows.append([
            c,
            _maybe_int(v["registros"]),
            _maybe_int(v["quantidade_total"]),
            part,
            "Ativa" if v["registros"] > 0 else "Sem movimento",
        ])

    _write_table(ws, 3, 1, ["Categoria", "Registros", "Qtd Total", "Participacao", "Status"], rows, max_rows=max(len(rows), 12), status_col_idx=4)
    ws.freeze_panes = "A4"
    if rows:
        ws.auto_filter.ref = f"A3:E{3 + len(rows)}"
    _autosize(ws, 5)


def _build_view_team(ws, by_team: Dict[str, dict]) -> None:
    ws.sheet_view.showGridLines = False
    _merge_box(ws, 1, 1, 1, 5, "SABESP - VISAO POR EQUIPE", fill=NAVY, color=WHITE, size=14, align="left")

    media = (sum(v["quantidade_total"] for v in by_team.values()) / max(1, len(by_team))) if by_team else 0
    rows = []
    for t, v in sorted(by_team.items()):
        qtd = float(v["quantidade_total"])
        if qtd == 0:
            leitura = "Sem producao"
        elif qtd >= media * 1.25:
            leitura = "Destaque"
        elif qtd < media * 0.75:
            leitura = "Monitorar"
        else:
            leitura = "Estavel"
        rows.append([
            t,
            v["nucleo"],
            _maybe_int(v["itens"]),
            _maybe_int(qtd),
            leitura,
        ])

    _write_table(ws, 3, 1, ["Equipe", "Nucleo", "Itens", "Qtd Total", "Leitura"], rows, max_rows=max(len(rows), 12), status_col_idx=4)
    ws.freeze_panes = "A4"
    if rows:
        ws.auto_filter.ref = f"A3:E{3 + len(rows)}"
    _autosize(ws, 5)


def build_management_workbook(
    output_dir: Path,
    dictionary_csv: Path | None = None,
    nucleo_reference_file: Path | None = None,
) -> Path:
    output_dir = Path(output_dir)

    exec_rows_raw = _filter_active_nucleo_rows(
        _reconcile_rows(_read_csv(output_dir / "execucao.csv"), nucleo_reference_file),
        nucleo_reference_file,
    )
    exec_rows = _countable_exec_rows(exec_rows_raw)
    occ_rows = _filter_active_nucleo_rows(
        _reconcile_rows(_read_csv(output_dir / "ocorrencias.csv"), nucleo_reference_file),
        nucleo_reference_file,
    )
    front_rows = _filter_active_nucleo_rows(
        _reconcile_rows(_read_csv(output_dir / "frentes.csv"), nucleo_reference_file),
        nucleo_reference_file,
    )
    dict_rows = _read_csv(Path(dictionary_csv)) if dictionary_csv and Path(dictionary_csv).exists() else []

    by_nucleus: Dict[str, dict] = defaultdict(lambda: {"frentes": 0, "itens": 0, "ocorrencias": 0, "quantidade_total": 0.0})
    by_category: Dict[str, dict] = defaultdict(lambda: {"registros": 0, "quantidade_total": 0.0})
    by_team: Dict[str, dict] = defaultdict(lambda: {"nucleo": "", "itens": 0, "quantidade_total": 0.0})
    by_occ_type: Counter = Counter()

    def _label_nucleo(nucleo: str) -> str:
        if not nucleo:
            return ""
        return nucleo

    for row in front_rows:
        nuc = str(row.get("nucleo", "") or "").strip()
        nuc_label = _label_nucleo(nuc)
        if nuc_label:
            by_nucleus[nuc_label]["frentes"] += 1

    for row in exec_rows:
        nuc = str(row.get("nucleo", "") or "").strip()
        nuc_label = _label_nucleo(nuc)
        cat = str(row.get("categoria_item", "") or "").strip()
        team = str(row.get("equipe", "") or "").strip()
        qty = _num(row.get("quantidade"))
        if nuc_label:
            by_nucleus[nuc_label]["itens"] += 1
            by_nucleus[nuc_label]["quantidade_total"] += qty
        if cat:
            by_category[cat]["registros"] += 1
            by_category[cat]["quantidade_total"] += qty
        if team:
            by_team[team]["nucleo"] = nuc_label or nuc
            by_team[team]["itens"] += 1
            by_team[team]["quantidade_total"] += qty

    for row in occ_rows:
        nuc = str(row.get("nucleo", "") or "").strip()
        nuc_label = _label_nucleo(nuc)
        tipo = str(row.get("tipo_ocorrencia", "") or "").strip()
        if nuc_label:
            by_nucleus[nuc_label]["ocorrencias"] += 1
        if tipo:
            by_occ_type[tipo] += 1

    frentes_sem_producao = sum(
        1
        for row in front_rows
        if str(row.get("status_frente", "") or "").strip().lower() in {"sem_producao", "paralisada"}
    )

    totals = {
        "nucleos": len([n for n, v in by_nucleus.items() if v["frentes"] or v["itens"] or v["ocorrencias"] or v["quantidade_total"]]),
        "frentes": len(front_rows),
        "ocorrencias": len(occ_rows),
        "quantidade_total": _maybe_int(sum(_num(row.get("quantidade")) for row in exec_rows)),
        "itens_execucao": len(exec_rows),
        "categorias_ativas": len([c for c, v in by_category.items() if v["registros"] > 0]),
        "equipes_ativas": len([k for k in by_team.keys() if k]),
    }

    ocorrencia_por_frente = round(totals["ocorrencias"] / max(1, totals["frentes"]), 2)
    if frentes_sem_producao > 0 or ocorrencia_por_frente >= 1.0:
        risco_operacional = "ALTO"
    elif totals["ocorrencias"] > 0:
        risco_operacional = "MODERADO"
    else:
        risco_operacional = "BAIXO"

    data_base = _first_non_empty(
        exec_rows[0].get("data_referencia") if exec_rows else "",
        occ_rows[0].get("data_referencia") if occ_rows else "",
        front_rows[0].get("data_referencia") if front_rows else "",
    )
    data_geracao = datetime.now().strftime("%d/%m/%Y %H:%M")

    wb = Workbook()
    ws_dash = wb.active
    ws_dash.title = "DASHBOARD_EXECUTIVO"
    ws_nuc = wb.create_sheet("VISAO_NUCLEOS")
    ws_cat = wb.create_sheet("VISAO_CATEGORIAS")
    ws_team = wb.create_sheet("VISAO_EQUIPES")
    ws_exec = wb.create_sheet("BASE_EXECUCAO")
    ws_occ = wb.create_sheet("BASE_OCORRENCIAS")
    ws_front = wb.create_sheet("BASE_FRENTES")
    ws_dic = wb.create_sheet("DICIONARIO_SERVICOS")
    ws_ins = wb.create_sheet("INSTRUCOES")

    _build_dashboard(
        ws_dash,
        totals=totals,
        by_nucleus=by_nucleus,
        by_category=by_category,
        by_team=by_team,
        by_occ_type=by_occ_type,
        frentes_sem_producao=frentes_sem_producao,
        ocorrencia_por_frente=ocorrencia_por_frente,
        risco_operacional=risco_operacional,
        data_base=data_base,
        data_geracao=data_geracao,
    )

    _build_view_nuc(ws_nuc, by_nucleus)
    _build_view_cat(ws_cat, by_category)
    _build_view_team(ws_team, by_team)

    _dump_base_sheet(
        ws_exec,
        "BASE EXECUCAO",
        exec_rows_raw,
        [
            "id_item",
            "id_frente",
            "data_referencia",
            "contrato",
            "programa",
            "nucleo_detectado_texto",
            "nucleo_oficial",
            "nucleo",
            "logradouro",
            "municipio_detectado_texto",
            "municipio_oficial",
            "municipio",
            "nucleo_status_cadastro",
            "equipe",
            "item_original",
            "item_normalizado",
            "categoria_item",
            "tipo_registro",
            "quantidade",
            "unidade",
            "servico_bruto",
            "servico_normalizado",
            "servico_oficial",
            "categoria",
            "material",
            "especificacao",
            "complemento",
            "observacao_item",
            "arquivo_origem",
        ],
    )

    _dump_base_sheet(
        ws_occ,
        "BASE OCORRENCIAS",
        occ_rows,
        [
            "id_ocorrencia",
            "id_frente",
            "data_referencia",
            "contrato",
            "programa",
            "nucleo_detectado_texto",
            "nucleo_oficial",
            "nucleo",
            "logradouro",
            "municipio_detectado_texto",
            "municipio_oficial",
            "municipio",
            "nucleo_status_cadastro",
            "equipe",
            "tipo_ocorrencia",
            "descricao",
            "impacto_producao",
            "arquivo_origem",
        ],
    )

    _dump_base_sheet(
        ws_front,
        "BASE FRENTES",
        front_rows,
        [
            "id_frente",
            "data_referencia",
            "contrato",
            "programa",
            "nucleo_detectado_texto",
            "nucleo_oficial",
            "nucleo",
            "logradouro",
            "municipio_detectado_texto",
            "municipio_oficial",
            "municipio",
            "nucleo_status_cadastro",
            "equipe",
            "status_frente",
            "observacao_frente",
            "arquivo_origem",
        ],
    )

    _dump_base_sheet(
        ws_dic,
        "DICIONARIO DE SERVICOS",
        dict_rows,
        ["nome_original", "nome_padronizado", "categoria", "unidade_padrao", "tipo_registro", "frase_tecnica", "alias"],
    )

    ws_ins.sheet_view.showGridLines = False
    _merge_box(ws_ins, 1, 1, 1, 8, "SABESP - GUIA DE USO DA BASE", fill=NAVY, color=WHITE, size=14, align="left")
    instrucoes = [
        "1. Esta base consolida os CSVs gerados pelo motor de processamento.",
        "2. A aba DASHBOARD_EXECUTIVO apresenta a leitura gerencial do periodo.",
        "3. As abas VISAO_* detalham desempenho por nucleo, categoria e equipe.",
        "4. As abas BASE_* preservam os registros operacionais para auditoria.",
        "5. A contagem de Caixa UMA evita dupla soma de itens de detalhamento (Embutida/Mureta).",
        "6. O risco operacional considera ocorrencias por frente e frentes sem producao.",
    ]
    for idx, texto in enumerate(instrucoes, start=3):
        c = ws_ins.cell(idx, 1, texto)
        _style_cell(c, fill=VERY_LIGHT if idx % 2 == 1 else WHITE, color=TEXT, align="left", wrap=True)
    ws_ins.column_dimensions["A"].width = 132

    output_path = output_dir / "base_gerencial.xlsx"
    wb.save(output_path)
    wb.close()
    return output_path









